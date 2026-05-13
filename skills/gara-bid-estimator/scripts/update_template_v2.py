#!/usr/bin/env python3
"""
update_template_v2.py — Aggiorna il template Excel `Stima_SolutionDesign_Template.xlsx` allo schema v2.0.

Modifiche:
1. Foglio `Requirements & Solution Mapping`:
   - Consolida legenda E/I/A nella sola cella N1 (libera O1 e P1)
   - Aggiunge header O3="Forbice & Driver" e P3="Rationale Stima"
   - Imposta larghezze e wrap text per O e P
2. Foglio `Summary`:
   - Inserisce riga 26 "Costi prodotti/servizi esterni" prima del TOTALE
   - Aggiorna le formule per includere la nuova riga (SUM ranges)
   - Sposta TOTALE/Contingency/Totale/Revenue di una riga
3. Nuovo foglio `Prodotti & Servizi Esterni` con header e footer TOTALE

NB: lo script è idempotente — se viene rieseguito, controlla lo stato attuale prima di modificare.
"""
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy

TEMPLATE = Path(r"C:\Users\g.sarno\.claude\skills\gara-bid-estimator\assets\Stima_SolutionDesign_Template.xlsx")

THIN = Side(border_style="thin", color="808080")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
HEADER_FILL = PatternFill(fill_type="solid", start_color="305496", end_color="305496")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_ALIGN_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
SUB_FILL = PatternFill(fill_type="solid", start_color="DCE6F1", end_color="DCE6F1")
TOTAL_FILL = PatternFill(fill_type="solid", start_color="FFC000", end_color="FFC000")
TOTAL_FONT = Font(name="Calibri", size=11, bold=True)


def update_requirements_sheet(wb):
    ws = wb["Requirements & Solution Mapping"]
    # 1. Consolida legenda E/I/A in N1, libera O1/P1
    if ws["N1"].value == "E=Estratto" or (ws["O1"].value == "I=Inferito"):
        ws["N1"] = "Legenda Certezza: E=Estratto | I=Inferito | A=Assunto"
        ws["O1"] = None
        ws["P1"] = None
    # 2. Header per O3 e P3 (se non già presenti)
    if ws["O3"].value != "Forbice & Driver":
        ws["O3"] = "Forbice & Driver"
        # Copia stile da N3 (header esistente)
        try:
            src = ws["N3"]
            ws["O3"].font = copy(src.font)
            ws["O3"].fill = copy(src.fill)
            ws["O3"].alignment = copy(src.alignment)
            ws["O3"].border = copy(src.border)
        except Exception:
            ws["O3"].font = HEADER_FONT
            ws["O3"].fill = HEADER_FILL
            ws["O3"].alignment = HEADER_ALIGN
            ws["O3"].border = BORDER
    if ws["P3"].value != "Rationale Stima":
        ws["P3"] = "Rationale Stima"
        try:
            src = ws["N3"]
            ws["P3"].font = copy(src.font)
            ws["P3"].fill = copy(src.fill)
            ws["P3"].alignment = copy(src.alignment)
            ws["P3"].border = copy(src.border)
        except Exception:
            ws["P3"].font = HEADER_FONT
            ws["P3"].fill = HEADER_FILL
            ws["P3"].alignment = HEADER_ALIGN
            ws["P3"].border = BORDER
    # 3. Larghezze colonne
    ws.column_dimensions["O"].width = 45
    ws.column_dimensions["P"].width = 35
    # 4. Wrap text nelle celle dati colonne O e P per righe 4-20
    for r in range(4, 21):
        for col in ("O", "P"):
            cell = ws[f"{col}{r}"]
            cell.alignment = DATA_ALIGN_WRAP
            # Copia bordo da L se presente
            try:
                src_border = ws[f"L{r}"].border
                if src_border:
                    cell.border = copy(src_border)
                else:
                    cell.border = BORDER
            except Exception:
                cell.border = BORDER
    print("[OK] Sheet 'Requirements & Solution Mapping': O/P aggiornate")


def update_summary_sheet(wb):
    ws = wb["Summary"]
    # Verifica se la riga 26 è già stata aggiornata (idempotenza)
    if ws["A26"].value == "Costi prodotti/servizi esterni":
        print("[SKIP] Summary già aggiornato")
        return

    # Stato atteso prima della modifica:
    # Riga 26: TOTALE PROGETTO | =SUM(B17:B25) | =IFERROR(D26/B26,0) | =SUM(D17:D25) | =SUM(E17:E25)
    # Riga 27: Contingency
    # Riga 28: Totale
    # Riga 29: Revenue
    assert ws["A26"].value == "TOTALE PROGETTO", f"A26 atteso 'TOTALE PROGETTO', trovato {ws['A26'].value!r}"

    # Salva stili e valori delle righe 26-29 prima della shift
    saved = {}
    for r in (26, 27, 28, 29):
        for c in range(1, 11):
            cell = ws.cell(r, c)
            saved[(r, c)] = {
                "value": cell.value,
                "font": copy(cell.font) if cell.font else None,
                "fill": copy(cell.fill) if cell.fill else None,
                "alignment": copy(cell.alignment) if cell.alignment else None,
                "border": copy(cell.border) if cell.border else None,
                "number_format": cell.number_format,
            }

    # Inserisce una riga vuota a indice 26 (riga 26 vecchia diventa 27, ecc.)
    ws.insert_rows(26)

    # Ripopola le righe 27-30 con i contenuti delle vecchie 26-29 + formule aggiornate
    # Formule da aggiornare: SUM(...:25) -> SUM(...:26); D26/B26 -> D27/B27; D26*20% -> D28*20%; SUM(D26:D27) -> SUM(D28:D29); D28*2 -> D30*2
    formula_remap = {
        "=SUM(B17:B25)": "=SUM(B17:B26)",
        "=SUM(D17:D25)": "=SUM(D17:D26)",
        "=SUM(E17:E25)": "=SUM(E17:E26)",
        "=IFERROR(D26/B26,0)": "=IFERROR(D27/B27,0)",
        "=D26*20%": "=D27*20%",
        "=SUM(D26:D27)": "=SUM(D27:D28)",
        "=D28*2": "=D29*2",
    }
    for r_old, r_new in [(26, 27), (27, 28), (28, 29), (29, 30)]:
        for c in range(1, 11):
            v = saved[(r_old, c)]["value"]
            new_val = formula_remap.get(v, v) if isinstance(v, str) else v
            cell = ws.cell(r_new, c)
            cell.value = new_val
            if saved[(r_old, c)]["font"] is not None:
                cell.font = saved[(r_old, c)]["font"]
            if saved[(r_old, c)]["fill"] is not None:
                cell.fill = saved[(r_old, c)]["fill"]
            if saved[(r_old, c)]["alignment"] is not None:
                cell.alignment = saved[(r_old, c)]["alignment"]
            if saved[(r_old, c)]["border"] is not None:
                cell.border = saved[(r_old, c)]["border"]
            cell.number_format = saved[(r_old, c)]["number_format"]

    # Pulizia: cancella i valori sulle vecchie celle 26-29 che ora sono "fantasma" (insert_rows li ha spostati ma openpyxl può lasciare residui sui style)
    # Riga 26 è stata appena inserita, quindi è già vuota a parte le celle che riempiamo sotto

    # Riempie la nuova riga 26 "Costi prodotti/servizi esterni"
    ws["A26"] = "Costi prodotti/servizi esterni"
    ws["B26"] = "NA"
    ws["C26"] = "NA"
    ws["D26"] = "=IFERROR(VLOOKUP(\"TOTALE\",'Prodotti & Servizi Esterni'!A:F,6,FALSE),0)"
    ws["E26"] = "=IFERROR(D26/SUM($D$17:$D$26),0)"

    # Aggiorna anche le formule % delle righe 17-25 da SUM($D$17:$D$25) a SUM($D$17:$D$26)
    for r in range(17, 26):
        cell = ws.cell(r, 5)  # colonna E
        if isinstance(cell.value, str) and "$D$17:$D$25" in cell.value:
            cell.value = cell.value.replace("$D$17:$D$25", "$D$17:$D$26")

    # Style per la nuova riga 26: replica stile della riga 25 (Test Performance)
    for c in range(1, 11):
        try:
            ref = ws.cell(25, c)
            tgt = ws.cell(26, c)
            if ref.font: tgt.font = copy(ref.font)
            if ref.fill: tgt.fill = copy(ref.fill)
            if ref.alignment: tgt.alignment = copy(ref.alignment)
            if ref.border: tgt.border = copy(ref.border)
            tgt.number_format = ref.number_format
        except Exception:
            pass

    print("[OK] Sheet 'Summary': aggiunta riga 26 'Costi prodotti/servizi esterni' e formule shift")


def create_prodotti_servizi_sheet(wb):
    sheet_name = "Prodotti & Servizi Esterni"
    if sheet_name in wb.sheetnames:
        print(f"[SKIP] Sheet '{sheet_name}' già esistente")
        return
    # Inserisce dopo "Requirements & Solution Mapping" (ordine voluto: Summary, RFP Analysis, Requirements, Prodotti & Servizi Esterni, ...)
    idx_req = wb.sheetnames.index("Requirements & Solution Mapping")
    ws = wb.create_sheet(sheet_name, idx_req + 1)

    # Title
    ws["A1"] = "PRODOTTI & SERVIZI ESTERNI — TECNOLOGIE NON IN SCOPE CUSTOM"
    ws.merge_cells("A1:I1")
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(fill_type="solid", start_color="305496", end_color="305496")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    ws["A2"] = ("Requisiti con categoria_scope != custom_software dal JSON v2.0. "
                "GG/U custom = 0; la colonna F è da compilare in fase commerciale con costi licenza/abbonamento/servizio.")
    ws.merge_cells("A2:I2")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="595959")
    ws.row_dimensions[2].height = 30

    # Headers (riga 3)
    headers = [
        "REQ ID",
        "Categoria",
        "Nome prodotto",
        "Fornitore",
        "Tipo costo",
        "Stima costo (€)",
        "Note",
        "Fonte bando",
        "Motivazione classificazione",
    ]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(3, i)
        c.value = h
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN
        c.border = BORDER
    ws.row_dimensions[3].height = 32

    # Righe placeholder vuote (4-13) per requisiti
    for r in range(4, 14):
        for c in range(1, 10):
            cell = ws.cell(r, c)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Riga TOTALE (14)
    ws["E14"] = "TOTALE"
    ws["E14"].font = TOTAL_FONT
    ws["E14"].fill = TOTAL_FILL
    ws["E14"].alignment = Alignment(horizontal="right", vertical="center")
    ws["E14"].border = BORDER
    ws["A14"] = "TOTALE"  # serve a VLOOKUP del Summary
    ws["A14"].font = TOTAL_FONT
    ws["A14"].fill = TOTAL_FILL
    ws["A14"].border = BORDER
    ws["F14"] = "=SUM(F4:F13)"
    ws["F14"].font = TOTAL_FONT
    ws["F14"].fill = TOTAL_FILL
    ws["F14"].border = BORDER
    ws["F14"].number_format = '#,##0.00 €'
    # Bordi per le altre celle della riga TOTALE
    for c in (2, 3, 4, 7, 8, 9):
        ws.cell(14, c).border = BORDER
        ws.cell(14, c).fill = TOTAL_FILL

    # Larghezze colonne
    widths = {"A": 12, "B": 18, "C": 25, "D": 18, "E": 22, "F": 16, "G": 35, "H": 22, "I": 35}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Formato numerico € per colonna F (righe 4-13)
    for r in range(4, 14):
        ws.cell(r, 6).number_format = '#,##0.00 €'

    print(f"[OK] Sheet '{sheet_name}' creato")


def main():
    if not TEMPLATE.exists():
        print(f"ERROR: template non trovato in {TEMPLATE}")
        sys.exit(1)
    wb = load_workbook(TEMPLATE)
    update_requirements_sheet(wb)
    update_summary_sheet(wb)
    create_prodotti_servizi_sheet(wb)
    wb.save(TEMPLATE)
    print(f"\n[DONE] Template salvato: {TEMPLATE}")


if __name__ == "__main__":
    main()
